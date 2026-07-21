"""Write approved answers back into the original questionnaire file.

For xlsx we open the original template, place the final_answer into the
detected answer column (or one column right of the question), and attach
citation entry ids as a cell comment. This preserves the customer's
formatting so the deliverable looks native.

For docx / pdf we produce a "companion" xlsx with question/answer columns
plus citations — round-tripping into a pdf template is out of scope for
the demo.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional

from app.models import Questionnaire, QuestionnaireItem


def export_questionnaire(
    q: Questionnaire,
    items: Iterable[QuestionnaireItem],
    output_dir: str | Path,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if q.file_kind == "xlsx":
        return _export_xlsx(q, list(items), output_dir)
    # docx / pdf → companion xlsx
    return _export_companion_xlsx(q, list(items), output_dir)


# ── xlsx export ───────────────────────────────────────────────────────


def _export_xlsx(q: Questionnaire, items: List[QuestionnaireItem], out_dir: Path) -> Path:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill

    wb = load_workbook(filename=q.original_path)
    meta = q.parse_meta or {}
    sheet_name = meta.get("sheet_name")
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    column_map = meta.get("column_map") or {}
    q_col = column_map.get("question")
    ans_col = column_map.get("answer") or meta.get("answer_col") or (q_col + 1 if q_col else 2)

    approved_fill = PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid")
    edited_fill = PatternFill(start_color="FEF3E2", end_color="FEF3E2", fill_type="solid")

    filled = 0
    for it in items:
        answer = _final(it)
        if not answer:
            continue
        cell = ws.cell(row=it.row_index, column=ans_col)
        cell.value = answer
        cell.fill = edited_fill if it.review_status == "edited" else approved_fill

        # Attach citations as a cell comment so reviewers see provenance in Excel.
        cite = _fmt_citations(it)
        if cite:
            cell.comment = Comment(cite, "Trust Copilot")
        filled += 1

    # Little audit banner in an unused cell far to the right.
    ws.cell(row=1, column=(ws.max_column or 1) + 2).value = (
        f"Filled by Trust Copilot — {filled} answers, "
        f"{q.total_items - filled} left for review."
    )

    out_path = out_dir / f"{q.id}_{Path(q.original_filename).stem}_filled.xlsx"
    wb.save(str(out_path))
    return out_path


# ── companion xlsx for docx / pdf inputs ──────────────────────────────


def _export_companion_xlsx(q: Questionnaire, items: List[QuestionnaireItem], out_dir: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws.append(["#", "Section", "Question", "Answer", "Confidence", "Status", "Citations"])
    ws.freeze_panes = "A2"

    for it in items:
        ws.append([
            it.row_index,
            it.section or "",
            it.question,
            _final(it) or "",
            round(it.confidence, 3),
            it.review_status,
            _fmt_citations(it) or "",
        ])

    # Rough column widths for readability.
    widths = [6, 24, 60, 60, 12, 14, 40]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = w

    out_path = out_dir / f"{q.id}_{Path(q.original_filename).stem}_answers.xlsx"
    wb.save(str(out_path))
    return out_path


# ── helpers ───────────────────────────────────────────────────────────


def _final(it: QuestionnaireItem) -> Optional[str]:
    """Prefer the reviewer's final_answer, fall back to the LLM draft."""
    if it.final_answer:
        return it.final_answer
    if it.drafted_answer and it.review_status in ("approved", "pending"):
        return it.drafted_answer
    return None


def _fmt_citations(it: QuestionnaireItem) -> Optional[str]:
    if not it.citation_entry_ids:
        return None
    ids = ", ".join(str(x)[:8] for x in it.citation_entry_ids)
    return f"Cited entries: {ids}\nConfidence: {it.confidence:.2f}"
